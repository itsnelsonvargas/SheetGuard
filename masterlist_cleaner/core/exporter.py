"""Excel workbook export with conditional formatting."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from sheetguard.models.results import DuplicateGroup, ProcessingResult, ValidationIssue

logger = logging.getLogger(__name__)

FILL_ERROR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_WARNING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_CORRECTED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FONT = Font(bold=True)


class WorkbookExporter:
    """Generate multi-sheet Excel reports from processing results."""

    def export_full_report(self, result: ProcessingResult, path: str | Path) -> Path:
        """Write CLEANED_DATA, VALIDATION_ERRORS, DUPLICATES, and SUMMARY sheets."""
        path = Path(path)
        wb = Workbook()
        wb.remove(wb.active)

        self._write_cleaned_sheet(wb, result)
        self._write_errors_sheet(wb, result.issues)
        self._write_duplicates_sheet(wb, result)
        self._write_summary_sheet(wb, result)

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        logger.info("Exported full report to %s", path)
        return path

    def export_cleaned_only(self, result: ProcessingResult, path: str | Path) -> Path:
        """Export only the cleaned data sheet."""
        path = Path(path)
        wb = Workbook()
        ws = wb.active
        ws.title = "CLEANED_DATA"
        self._populate_dataframe(ws, result.cleaned_df)
        self._apply_issue_formatting(ws, result)
        wb.save(path)
        return path

    def export_validation_report(self, result: ProcessingResult, path: str | Path) -> Path:
        path = Path(path)
        wb = Workbook()
        wb.remove(wb.active)
        self._write_errors_sheet(wb, result.issues)
        self._write_summary_sheet(wb, result)
        wb.save(path)
        return path

    def export_duplicate_report(
        self, result: ProcessingResult, dup_df: pd.DataFrame, path: str | Path
    ) -> Path:
        path = Path(path)
        wb = Workbook()
        ws = wb.active
        ws.title = "DUPLICATES"
        self._populate_dataframe(ws, dup_df)
        wb.save(path)
        return path

    def _write_cleaned_sheet(self, wb: Workbook, result: ProcessingResult) -> None:
        ws = wb.create_sheet("CLEANED_DATA")
        self._populate_dataframe(ws, result.cleaned_df)
        self._apply_issue_formatting(ws, result)

    def _write_errors_sheet(self, wb: Workbook, issues: list[ValidationIssue]) -> None:
        ws = wb.create_sheet("VALIDATION_ERRORS")
        headers = [
            "row",
            "field_id",
            "column",
            "severity",
            "rule_type",
            "message",
            "original_value",
            "cleaned_value",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
        for issue in issues:
            ws.append(
                [
                    issue.row_index + 1,
                    issue.field_id,
                    issue.column,
                    issue.severity,
                    issue.rule_type,
                    issue.message,
                    issue.original_value,
                    issue.cleaned_value,
                ]
            )
            row_num = ws.max_row
            fill = FILL_WARNING if issue.severity == "warning" else FILL_ERROR
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
        self._autosize_columns(ws)

    def _write_duplicates_sheet(self, wb: Workbook, result: ProcessingResult) -> None:
        ws = wb.create_sheet("DUPLICATES")
        if not result.duplicates:
            ws.append(["No duplicates found"])
            return
        from sheetguard.core.duplicate_checker import DuplicateChecker

        checker = DuplicateChecker.__new__(DuplicateChecker)
        dup_df = self._duplicates_df(result)
        self._populate_dataframe(ws, dup_df)

    def _duplicates_df(self, result: ProcessingResult) -> pd.DataFrame:
        rows: list[dict[str, Any]] = []
        for g in result.duplicates:
            for row_idx in g.row_indices:
                row = result.cleaned_df.iloc[row_idx].to_dict()
                row["duplicate_rule"] = g.rule_name
                row["row_number"] = row_idx + 1
                row["duplicate_key"] = str(g.key_values)
                rows.append(row)
        return pd.DataFrame(rows) if rows else pd.DataFrame()

    def _write_summary_sheet(self, wb: Workbook, result: ProcessingResult) -> None:
        ws = wb.create_sheet("SUMMARY")
        summary = result.summary or {}
        metrics = [
            ("Total Rows", summary.get("total_rows", len(result.cleaned_df))),
            ("Errors", result.error_count),
            ("Warnings", result.warning_count),
            ("Corrections", len(result.corrections)),
            ("Duplicate Groups", len(result.duplicates)),
        ]
        ws.append(["Metric", "Value"])
        for cell in ws[1]:
            cell.font = HEADER_FONT
        for name, val in metrics:
            ws.append([name, val])
        ws.append([])
        ws.append(["Duplicate Rules Breakdown"])
        by_rule: dict[str, int] = {}
        for g in result.duplicates:
            by_rule[g.rule_name] = by_rule.get(g.rule_name, 0) + 1
        for rule, count in by_rule.items():
            ws.append([rule, count])
        self._autosize_columns(ws)

    def _populate_dataframe(self, ws: Any, df: pd.DataFrame) -> None:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.font = HEADER_FONT
        self._autosize_columns(ws)

    def _apply_issue_formatting(self, ws: Any, result: ProcessingResult) -> None:
        col_index = {name: i + 1 for i, name in enumerate(result.cleaned_df.columns)}
        for issue in result.issues:
            col = col_index.get(issue.column)
            if col is None:
                continue
            row = issue.row_index + 2
            cell = ws.cell(row=row, column=col)
            if issue.severity == "warning":
                cell.fill = FILL_WARNING
            else:
                cell.fill = FILL_ERROR
        for (row_idx, col_name), _ in result.corrections.items():
            col = col_index.get(col_name)
            if col is None:
                continue
            ws.cell(row=row_idx + 2, column=col).fill = FILL_CORRECTED

    @staticmethod
    def _autosize_columns(ws: Any, max_width: int = 50) -> None:
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            length = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[letter].width = min(length + 2, max_width)
